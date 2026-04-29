import csv
import io
from django.contrib import admin
from django.http import HttpResponse
from .models import Category, Collection, Product, ProductImage, ProductSizeInventory, Cart, CartItem, Order, OrderItem, ContactMessage, NewsletterSubscriber


class ProductSizeInventoryInline(admin.TabularInline):
    model = ProductSizeInventory
    extra = 0
    fields = ['size', 'quantity']
    readonly_fields = ['size']
    can_delete = False

    def has_add_permission(self, request, obj=None):
        return False


class ProductImageInline(admin.TabularInline):
    model = ProductImage
    extra = 1
    fields = ['image', 'image_url', 'alt_text', 'is_primary', 'preview']
    readonly_fields = ['preview']

    def preview(self, obj):
        from django.utils.html import format_html
        if obj.image:
            return format_html('<img src="{}" style="height:80px;object-fit:cover;border-radius:4px;">', obj.image.url)
        if obj.image_url:
            return format_html('<img src="{}" style="height:80px;object-fit:cover;border-radius:4px;">', obj.image_url)
        return '-'
    preview.short_description = 'Preview'


class CartItemInline(admin.TabularInline):
    model = CartItem
    extra = 0
    readonly_fields = ['get_subtotal']


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ['get_subtotal']


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ['name', 'slug']
    prepopulated_fields = {'slug': ('name',)}


@admin.register(Collection)
class CollectionAdmin(admin.ModelAdmin):
    list_display = ['name', 'slug']
    prepopulated_fields = {'slug': ('name',)}


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ['name', 'category', 'collection', 'price', 'compare_price', 'in_stock', 'featured']
    list_filter = ['category', 'collection', 'in_stock', 'featured']
    search_fields = ['name', 'description']
    prepopulated_fields = {'slug': ('name',)}
    inlines = [ProductSizeInventoryInline, ProductImageInline]

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        for size in ['XS', 'S', 'M', 'L', 'XL', 'XXL']:
            ProductSizeInventory.objects.get_or_create(product=obj, size=size, defaults={'quantity': 0})


@admin.register(Cart)
class CartAdmin(admin.ModelAdmin):
    list_display = ['cart_token', 'user', 'created_at', 'get_item_count', 'get_total']
    readonly_fields = ['cart_token', 'created_at', 'updated_at']
    inlines = [CartItemInline]


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ['id', 'first_name', 'last_name', 'email', 'total', 'status', 'created_at', 'delete_button']
    list_filter = ['status']
    search_fields = ['email', 'first_name', 'last_name']
    inlines = [OrderItemInline]

    def delete_button(self, obj):
        from django.utils.html import format_html
        from django.urls import reverse
        url = reverse('admin:djangoapp_order_delete', args=[obj.pk])
        return format_html(
            '<a href="{}" style="color:#dc2626;font-weight:600;">Delete</a>', url
        )
    delete_button.short_description = ''


@admin.register(NewsletterSubscriber)
class NewsletterSubscriberAdmin(admin.ModelAdmin):
    list_display = ['email', 'joined_at']
    search_fields = ['email']
    readonly_fields = ['joined_at']
    actions = ['export_csv', 'export_excel']

    def export_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="newsletter_subscribers.csv"'
        writer = csv.writer(response)
        writer.writerow(['Email', 'Joined At'])
        for sub in queryset:
            writer.writerow([sub.email, sub.joined_at.strftime('%Y-%m-%d %H:%M:%S')])
        return response
    export_csv.short_description = 'Download selected as CSV'

    def export_excel(self, request, queryset):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Subscribers'

        header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 22

        for col, heading in enumerate(['Email', 'Joined At'], start=1):
            cell = ws.cell(row=1, column=col, value=heading)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row, sub in enumerate(queryset, start=2):
            ws.cell(row=row, column=1, value=sub.email)
            ws.cell(row=row, column=2, value=sub.joined_at.strftime('%Y-%m-%d %H:%M:%S'))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="newsletter_subscribers.xlsx"'
        return response
    export_excel.short_description = 'Download selected as Excel'


@admin.register(ContactMessage)
class ContactMessageAdmin(admin.ModelAdmin):
    list_display = ['name', 'email', 'subject', 'created_at', 'is_read']
    list_filter = ['is_read']
    search_fields = ['name', 'email', 'subject']
